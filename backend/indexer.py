import os
import re
import csv
import logging
from typing import List, Dict, Any, Tuple
import pypdf
import docx
import openpyxl
from sentence_transformers import SentenceTransformer
from backend.config import settings, logger
from backend.models import OpsBrainException

class DocumentIndexer:
    _model = None

    @classmethod
    def get_model(cls) -> SentenceTransformer:
        if cls._model is None:
            logger.info(f"Loading local embedding model: {settings.EMBEDDING_MODEL_NAME}...")
            try:
                cls._model = SentenceTransformer(settings.EMBEDDING_MODEL_NAME)
                logger.info("Local embedding model loaded successfully!")
            except Exception as e:
                logger.error(f"Failed to load embedding model: {e}")
                raise OpsBrainException(f"Failed to initialize embedding engine: {e}", code="EMBEDDING_INIT_FAILED")
        return cls._model

    def extract_text_and_metadata(self, filepath: str, file_type: str) -> Tuple[str, Dict[str, Any]]:
        if not os.path.exists(filepath):
            raise OpsBrainException(f"File not found: {filepath}", code="FILE_NOT_FOUND", status_code=404)

        text = ""
        metadata = {
            "filename": os.path.basename(filepath),
            "file_size_bytes": os.path.getsize(filepath),
            "file_type": file_type
        }

        try:
            if file_type == "PDF":
                reader = pypdf.PdfReader(filepath)
                pages_text = []
                for i, page in enumerate(reader.pages):
                    page_text = page.extract_text()
                    if page_text:
                        pages_text.append(page_text)
                text = "\n".join(pages_text)
                metadata["page_count"] = len(reader.pages)

            elif file_type == "DOCX":
                doc = docx.Document(filepath)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                text = "\n".join(paragraphs)
                metadata["paragraph_count"] = len(paragraphs)

            elif file_type == "XLSX":
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheets_text = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    sheet_rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_vals = [str(val) for val in row if val is not None]
                        if row_vals:
                            sheet_rows.append(" | ".join(row_vals))
                    if sheet_rows:
                        sheets_text.append(f"--- Sheet: {sheet} ---\n" + "\n".join(sheet_rows))
                text = "\n\n".join(sheets_text)
                metadata["sheets"] = wb.sheetnames

            elif file_type == "CSV":
                rows_text = []
                with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        row_vals = [str(val) for val in row if val is not None]
                        if row_vals:
                            rows_text.append(" | ".join(row_vals))
                text = "\n".join(rows_text)
                metadata["row_count"] = len(rows_text)

            elif file_type == "TXT":
                with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read()

            else:
                raise OpsBrainException(f"Unsupported file type: {file_type}", code="UNSUPPORTED_FILE_TYPE", status_code=400)

            return text, metadata

        except OpsBrainException:
            raise
        except Exception as e:
            logger.error(f"Error extracting text from file {filepath} ({file_type}): {e}")
            raise OpsBrainException(f"Text extraction failed: {str(e)}", code="TEXT_EXTRACTION_FAILED")

    def chunk_text(self, text: str, metadata: Dict[str, Any], chunk_size: int = 500, overlap: int = 50) -> List[Dict[str, Any]]:
        # Slide text into semantic chunks
        if not text.strip():
            return []

        chunks = []
        filename = metadata.get("filename", "Unknown")
        file_type = metadata.get("file_type", "Unknown")
        
        # Simple split by paragraphs/newlines first
        paragraphs = text.split("\n")
        current_chunk = []
        current_len = 0
        page_num = 1

        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
                
            para_len = len(para)
            if current_len + para_len > chunk_size and current_chunk:
                # Compile chunk
                chunk_text = "\n".join(current_chunk)
                
                # Context Header Injection (Cerebro style)
                header = f"/* SOURCE: {filename} | TYPE: {file_type} */\n"
                final_content = header + chunk_text
                
                chunks.append({
                    "content": final_content,
                    "page_number": page_num
                })
                
                # Maintain overlap (last paragraph)
                overlap_p = current_chunk[-1] if len(current_chunk) > 1 else ""
                current_chunk = [overlap_p, para] if overlap_p else [para]
                current_len = len(overlap_p) + para_len
            else:
                current_chunk.append(para)
                current_len += para_len
                
        # Append residual chunk
        if current_chunk:
            chunk_text = "\n".join(current_chunk)
            header = f"/* SOURCE: {filename} | TYPE: {file_type} */\n"
            chunks.append({
                "content": header + chunk_text,
                "page_number": page_num
            })

        logger.info(f"Split document into {len(chunks)} chunks.")
        return chunks

    def generate_embeddings(self, chunks: List[str]) -> List[List[float]]:
        model = self.get_model()
        try:
            # Local model encoding runs on CPU, returns numpy array, cast to list of floats
            embeddings = model.encode(chunks, convert_to_numpy=True).tolist()
            return embeddings
        except Exception as e:
            logger.error(f"Failed to generate embeddings: {e}")
            raise OpsBrainException(f"Embedding generation failed: {e}", code="EMBEDDING_GENERATION_FAILED")
